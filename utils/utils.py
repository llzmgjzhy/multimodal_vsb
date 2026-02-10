import torch
import numpy as np
import matplotlib.pyplot as plt
from sklearn.model_selection import StratifiedKFold
import os
import builtins
import sys
import logging
from openpyxl import Workbook, load_workbook
import pandas as pd
from sklearn.metrics import (
    precision_recall_fscore_support,
    accuracy_score,
    roc_auc_score,
)

logging.basicConfig(
    format="%(asctime)s | %(levelname)s : %(message)s", level=logging.INFO
)
logger = logging.getLogger("__main__")


def matthews_correlation(y_true, y_pred):
    """Calculates the Matthews correlation coefficient measure for quality of binary classification problems."""
    y_pred = torch.tensor(y_pred, dtype=torch.float32)
    y_true = torch.tensor(y_true, dtype=torch.float32)

    y_pred_pos = torch.round(torch.clamp(y_pred, 0, 1))
    y_pred_neg = 1 - y_pred_pos

    y_pos = torch.round(torch.clamp(y_true, 0, 1))
    y_neg = 1 - y_pos

    tp = torch.sum(y_pos * y_pred_pos)
    tn = torch.sum(y_neg * y_pred_neg)

    fp = torch.sum(y_neg * y_pred_pos)
    fn = torch.sum(y_pos * y_pred_neg)

    numerator = tp * tn - fp * fn
    denominator = torch.sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))

    return numerator / (denominator + torch.finfo(torch.float32).eps)


def mcc(tp, tn, fp, fn):
    sup = tp * tn - fp * fn
    inf = (tp + fp) * (tp + fn) * (tn + fp) * (tn + fn)
    if inf == 0:
        return 0
    else:
        return sup / np.sqrt(inf)


def eval_mcc(y_true, y_prob, show=False):
    """
    A fast implementation of Anokas mcc optimization code.

    This code takes as input probabilities, and selects the threshold that
    yields the best MCC score. It is efficient enough to be used as a
    custom evaluation function in xgboost

    Source: https://www.kaggle.com/cpmpml/optimizing-probabilities-for-best-mcc
    Source: https://www.kaggle.com/c/bosch-production-line-performance/forums/t/22917/optimising-probabilities-binary-prediction-script
    Creator: CPMP
    """
    idx = np.argsort(y_prob)
    y_true_sort = y_true[idx]
    n = y_true.shape[0]
    nump = 1.0 * np.sum(y_true)  # number of positive
    numn = n - nump  # number of negative
    tp = nump
    tn = 0.0
    fp = numn
    fn = 0.0
    best_mcc = 0.0
    best_id = -1
    prev_proba = -1
    best_proba = -1
    mccs = np.zeros(n)
    for i in range(n):
        # all items with idx < i are predicted negative while others are predicted positive
        # only evaluate mcc when probability changes
        proba = y_prob[idx[i]]
        if proba != prev_proba:
            new_mcc = mcc(tp, tn, fp, fn)
            if new_mcc >= best_mcc:
                best_mcc = new_mcc
                best_id = i
                best_proba = (prev_proba + proba) / 2.0 if prev_proba >= 0 else proba
            prev_proba = proba
        mccs[i] = new_mcc
        if y_true_sort[i] == 1:
            tp -= 1.0
            fn += 1.0
        else:
            fp -= 1.0
            tn += 1.0
    if show:
        y_pred = (y_prob >= best_proba).astype(int)
        score = matthews_correlation(y_true, y_pred)
        plt.plot(mccs)
        return best_proba, best_mcc, y_pred
    else:
        return best_proba, best_mcc, None


def stratified_train_val_test_split(X, y, num_folds=5, seed=42):
    np.random.seed(seed)
    skf = StratifiedKFold(n_splits=num_folds, shuffle=True, random_state=seed)
    # get the indices of the folds
    splits = np.zeros_like(y, dtype=int)
    for fold_id, (_, val_idx) in enumerate(skf.split(X, y)):
        splits[val_idx] = fold_id

    fold_splits = []
    for val_fold in range(num_folds):
        test_fold = (val_fold + 1) % num_folds
        train_folds = [f for f in range(num_folds) if f not in [val_fold, test_fold]]

        train_idx = np.where(np.isin(splits, train_folds))[0]
        val_idx = np.where(splits == val_fold)[0]
        test_idx = np.where(splits == test_fold)[0]

        fold_splits.append((train_idx, val_idx, test_idx))

    return fold_splits


def create_dirs(dirs):
    """
    Input:
        dirs: a list of directories to create, in case these directories are not found
    Returns:
        exit_code: 0 if success, -1 if failure
    """
    try:
        for dir_ in dirs:
            if not os.path.exists(dir_):
                os.makedirs(dir_)
        return 0
    except Exception as err:
        print("Creating directories error: {0}".format(err))
        exit(-1)


class Printer(object):
    """Class for printing output by refreshing the same line in the console, e.g. for indicating progress of a process"""

    def __init__(self, console=True):

        if console:
            self.print = self.dyn_print
        else:
            self.print = builtins.print

    @staticmethod
    def dyn_print(data):
        """Print things to stdout on one line, refreshing it dynamically"""
        sys.stdout.write("\r\x1b[K" + data.__str__())
        sys.stdout.flush()


def readable_time(time_difference):
    """Convert a float measuring time difference in seconds into a tuple of (hours, minutes, seconds)"""

    hours = time_difference // 3600
    minutes = (time_difference // 60) % 60
    seconds = time_difference % 60

    return hours, minutes, seconds


def export_performance_metrics(
    filepath, metrics_table, header, book=None, sheet_name="metrics"
):
    """Exports performance metrics on the validation set for all epochs to an excel file"""

    if os.path.exists(filepath):  # Create a records file for the first time
        book = load_workbook(filepath)

    if book is None:
        book = Workbook()  # new excel work book
        del book["Sheet"]  # remove default sheet

    book = write_table_to_sheet([header] + metrics_table, book, sheet_name=sheet_name)

    book.save(filepath)
    logger.info("Exported per epoch performance metrics in '{}'".format(filepath))

    return book


def write_table_to_sheet(table, work_book, sheet_name=None):
    """Writes a table implemented as a list of lists to an excel sheet in the given work book object"""

    sheet = work_book.create_sheet(sheet_name, index=0)

    for row_ind, row_list in enumerate(table):
        write_row(sheet, row_ind, row_list)

    return work_book


def write_row(sheet, row_ind, data_list):
    """Write a list to row_ind row of an excel sheet"""

    # row = sheet.row(row_ind)
    for col_ind, col_value in enumerate(data_list, start=1):
        sheet.cell(
            row=row_ind + 1, column=col_ind, value=col_value
        )  # row and col starts from 1 in openpyxl
    return


def save_model(path, epoch, model, optimizer=None):
    if isinstance(model, torch.nn.DataParallel):
        state_dict = model.module.state_dict()
    else:
        state_dict = model.state_dict()
    data = {"epoch": epoch, "state_dict": state_dict}
    if not (optimizer is None):
        data["optimizer"] = optimizer.state_dict()
    torch.save(data, path)


def export_record(filepath, values):
    """Adds a list of values as a bottom row of a table in a given excel file"""

    # read_book = xlrd.open_workbook(filepath, formatting_info=True)
    read_book = load_workbook(filepath)
    sheet = read_book.active
    last_row = sheet.max_row

    write_row(sheet, last_row, values)
    read_book.save(filepath)


def register_test_record(
    filepath,
    timestamp,
    experiment_name,
    best_metrics,
    final_metrics=None,
    test_metrics=None,
    comment="",
):
    """
    Adds the best and final metrics of a given experiment as a record in an excel sheet with other experiment records.
    Creates excel sheet if it doesn't exist.
    Args:
        filepath: path of excel file keeping records
        timestamp: string
        experiment_name: string
        best_metrics: dict of metrics at best epoch {metric_name: metric_value}. Includes "epoch" as first key
        final_metrics: dict of metrics at final epoch {metric_name: metric_value}. Includes "epoch" as first key
        test_metrics: dict of metrics at test stage {metric_name: metric_value}.
        comment: optional description
    """
    metrics_names, metrics_values = zip(*best_metrics.items())
    row_values = [timestamp, experiment_name, comment] + list(metrics_values)
    if final_metrics is not None:
        final_metrics_names, final_metrics_values = zip(*final_metrics.items())
        row_values += list(final_metrics_values)

    if test_metrics is not None:
        test_metrics_names, test_metrics_values = zip(*test_metrics.items())
        row_values += list(test_metrics_values)

    if not os.path.exists(filepath):  # Create a records file for the first time
        logger.warning(
            "Records file '{}' does not exist! Creating new file ...".format(filepath)
        )
        directory = os.path.dirname(filepath)
        if len(directory) and not os.path.exists(directory):
            os.makedirs(directory)
        header = ["Timestamp", "Name", "Comment"] + ["Best " + m for m in metrics_names]
        if final_metrics is not None:
            header += ["Final " + m for m in final_metrics_names]
        if test_metrics is not None:
            header += [m for m in test_metrics_names]
        book = Workbook()  # new excel work book
        del book["Sheet"]  # excel work book
        book = write_table_to_sheet([header, row_values], book, sheet_name="records")
        book.save(filepath)
    else:
        try:
            export_record(filepath, row_values)
        except Exception as x:
            alt_path = os.path.join(
                os.path.dirname(filepath), "record_" + experiment_name
            )
            logger.error(
                "Failed saving in: '{}'! Will save here instead: {}".format(
                    filepath, alt_path
                )
            )
            export_record(alt_path, row_values)
            filepath = alt_path

    logger.info("Exported performance record to '{}'".format(filepath))


# def itr_test_result(config):
#     """
#     Calculate one iteration's test result for anomaly detection.
#     """

#     def load_fold_results(prefix):
#         probs, labels = [], []
#         for i in range(config.split_num):
#             path = os.path.join(config.pred_dir, f"{prefix}_pred_{i}.csv")
#             df = pd.read_csv(path)
#             probs.extend(df["pred"].values)
#             labels.extend(df["targets"].values)
#         return np.array(probs), np.array(labels)

#     all_val_probs, all_val_labels = load_fold_results("val")
#     all_test_probs, all_test_labels = load_fold_results("test")

#     # get the best threshold
#     best_threshold, best_val_mcc, _ = eval_mcc(all_val_labels, all_val_probs)
#     logger.info(f"Best threshold: {best_threshold}, Best val MCC: {best_val_mcc:.4f}")

#     # get the test mcc
#     all_test_pred = (all_test_probs > best_threshold).astype(int)
#     test_mcc = matthews_correlation(all_test_labels, all_test_pred).item()
#     test_accuracy = accuracy_score(all_test_labels, all_test_pred)
#     precision, recall, f_score, support = precision_recall_fscore_support(
#         all_test_labels, all_test_pred, average="binary"
#     )
#     test_auc = roc_auc_score(all_test_labels, all_test_probs)

#     logger.info(f"Test MCC: {test_mcc:.4f}")

#     result = {
#         "best_threshold": best_threshold,
#         "best_val_mcc": best_val_mcc,
#         "test_mcc": test_mcc,
#         "test_accuracy": test_accuracy,
#         "test_precision": precision,
#         "test_recall": recall,
#         "test_f_score": f_score,
#         "test_auc": test_auc,
#     }

#     # save csv with test results
#     test_results_df = pd.DataFrame([result])
#     test_results_df.to_csv(
#         os.path.join(config.pred_dir, "test_results.csv"), index=False
#     )

#     return result


def itr_test_result(config):
    """
    Calculate one iteration's test result for anomaly detection.

    New features:
    - Per-fold: find best threshold on that fold's val set, then evaluate on that fold's test set.
    - Save per-fold thresholds and metrics.
    - Keep previous global-merged functionality.
    """

    def load_one_fold(prefix: str, fold_idx: int):
        path = os.path.join(config.pred_dir, f"{prefix}_pred_{fold_idx}.csv")
        df = pd.read_csv(path)

        # safety checks
        if "pred" not in df.columns or "targets" not in df.columns:
            raise ValueError(
                f"Missing required columns in {path}. "
                f"Found columns: {list(df.columns)}; required: ['pred','targets']"
            )

        probs = df["pred"].values.astype(float)
        labels = df["targets"].values.astype(int)
        sample_ids = df["sample_id"].values.astype(int)
        return probs, labels, sample_ids

    def load_all_folds(prefix: str):
        probs_all, labels_all, sample_ids_all = [], [], []
        for i in range(config.split_num):
            probs, labels, sample_ids = load_one_fold(prefix, i)
            probs_all.append(probs)
            labels_all.append(labels)
            sample_ids_all.append(sample_ids)
        return (
            np.concatenate(probs_all),
            np.concatenate(labels_all),
            np.concatenate(sample_ids_all),
        )

    def compute_metrics(y_true: np.ndarray, probs: np.ndarray, threshold: float):
        y_pred = (probs > threshold).astype(int)

        mcc = matthews_correlation(y_true, y_pred).item()
        acc = accuracy_score(y_true, y_pred)

        precision, recall, f_score, _ = precision_recall_fscore_support(
            y_true, y_pred, average="binary", zero_division=0
        )

        # AUC requires both classes present; otherwise sklearn raises.
        try:
            auc = roc_auc_score(y_true, probs)
        except ValueError:
            auc = np.nan

        return {
            "mcc": mcc,
            "accuracy": acc,
            "precision": precision,
            "recall": recall,
            "f_score": f_score,
            "auc": auc,
        }

    def _build_merged_df(meta_data_train, sample_ids, preds):
        id_meas = meta_data_train["id_measurement"].drop_duplicates().to_numpy()
        id_meas_sample = id_meas[sample_ids]
        yp_df = pd.DataFrame(
            {
                "id_measurement": id_meas_sample.astype(int),
                "prediction": np.asarray(preds, dtype=float).reshape(-1),
            }
        )

        expanded = meta_data_train[["id_measurement", "signal_id", "target"]].copy()
        expanded["id_measurement"] = expanded["id_measurement"].astype(int)

        merged_df = expanded.merge(yp_df, on="id_measurement", how="inner")
        return merged_df

    # -------------------------
    # 1) Per-fold evaluation
    # -------------------------
    fold_rows = []
    for fold in range(config.split_num):
        val_probs, val_labels, _ = load_one_fold("val", fold)
        test_probs, test_labels, _ = load_one_fold("test", fold)

        # Find best threshold on THIS fold's val set
        best_th, best_val_mcc, _ = eval_mcc(val_labels, val_probs)

        # Evaluate test with this fold-specific threshold
        test_metrics = compute_metrics(test_labels, test_probs, best_th)

        fold_rows.append(
            {
                "fold": fold,
                "best_threshold": float(best_th),
                "best_val_mcc": float(best_val_mcc),
                "test_mcc": float(test_metrics["mcc"]),
                "test_accuracy": float(test_metrics["accuracy"]),
                "test_precision": float(test_metrics["precision"]),
                "test_recall": float(test_metrics["recall"]),
                "test_f_score": float(test_metrics["f_score"]),
                "test_auc": (
                    float(test_metrics["auc"])
                    if not np.isnan(test_metrics["auc"])
                    else np.nan
                ),
                "val_size": int(len(val_labels)),
                "test_size": int(len(test_labels)),
                "val_pos_rate": float(np.mean(val_labels)),
                "test_pos_rate": float(np.mean(test_labels)),
            }
        )

    fold_df = pd.DataFrame(fold_rows)

    # summary across folds (mean + std)
    per_fold_test_mcc_mean = float(fold_df["test_mcc"].mean())
    per_fold_test_mcc_std = (
        float(fold_df["test_mcc"].std(ddof=1)) if len(fold_df) > 1 else 0.0
    )

    per_fold_threshold_mean = float(fold_df["best_threshold"].mean())
    per_fold_threshold_std = (
        float(fold_df["best_threshold"].std(ddof=1)) if len(fold_df) > 1 else 0.0
    )

    # Save per-fold details
    fold_path = os.path.join(config.pred_dir, "fold_results.csv")
    fold_df.to_csv(fold_path, index=False)

    # -------------------------
    # 2) Global merged evaluation (measurment-level)
    # -------------------------
    all_val_probs_meas, all_val_labels, all_val_sample_ids = load_all_folds("val")
    all_test_probs_meas, all_test_labels, all_test_sample_ids = load_all_folds("test")

    global_best_th_meas, global_best_val_mcc_meas, _ = eval_mcc(
        all_val_labels, all_val_probs_meas
    )

    global_test_metrics_meas = compute_metrics(
        all_test_labels, all_test_probs_meas, global_best_th_meas
    )

    # -------------------------
    # 3) Global merged evaluation (signal-level)
    # -------------------------
    meta_data_train = pd.read_csv(
        os.path.join(config.root_path, "VSBdata", "metadata_train.csv")
    )
    val_exp = _build_merged_df(meta_data_train, all_val_sample_ids, all_val_probs_meas)
    test_exp = _build_merged_df(
        meta_data_train, all_test_sample_ids, all_test_probs_meas
    )
    global_best_th_exp, global_best_val_mcc_exp, _ = eval_mcc(
        val_exp["target"].values.astype(float),  # signal-level y
        val_exp["prediction"].values.astype(float),  # broadcasted measurement score
    )
    global_exp_test_pred = (
        test_exp["prediction"].values.astype(float) > global_best_th_exp
    ).astype(int)
    global_exp_test_mcc = matthews_correlation(
        test_exp["target"].values.astype(int), global_exp_test_pred
    ).item()

    # -------------------------
    # 4) Logging + save summary
    # -------------------------
    logger.info(
        f"[Per-fold] test MCC mean±std: {per_fold_test_mcc_mean:.4f} ± {per_fold_test_mcc_std:.4f} | "
        f"threshold mean±std: {per_fold_threshold_mean:.6f} ± {per_fold_threshold_std:.6f}"
    )
    logger.info(
        f"[Global] Best threshold (measurement): {global_best_th_meas:.6f}, Best val MCC (measurement): {global_best_val_mcc_meas:.4f}, "
        f"Test MCC (measurement): {global_test_metrics_meas['mcc']:.4f}"
    )
    logger.info(
        f"[Global] Best threshold: {global_best_th_exp:.6f}, "
        f"Best val MCC: {global_best_val_mcc_exp:.4f}, "
        f"Test MCC: {global_exp_test_mcc:.4f}"
    )
    logger.info(f"Per-fold details saved to: {fold_path}")

    result = {
        # ---- global (measurement) ----
        "global_best_threshold (measurement)": float(global_best_th_meas),
        "global_best_val_mcc (measurement)": float(global_best_val_mcc_meas),
        "global_test_mcc (measurement)": float(global_test_metrics_meas["mcc"]),
        "global_test_accuracy (measurement)": float(
            global_test_metrics_meas["accuracy"]
        ),
        "global_test_precision (measurement)": float(
            global_test_metrics_meas["precision"]
        ),
        "global_test_recall (measurement)": float(global_test_metrics_meas["recall"]),
        "global_test_f_score (measurement)": float(global_test_metrics_meas["f_score"]),
        "global_test_auc (measurement)": (
            float(global_test_metrics_meas["auc"])
            if not np.isnan(global_test_metrics_meas["auc"])
            else np.nan
        ),
        # ---- global (signal-level) ----
        "global_best_threshold": float(global_best_th_exp),
        "global_best_val_mcc": float(global_best_val_mcc_exp),
        "global_exp_test_mcc": float(global_exp_test_mcc),
        # ---- per-fold summary ----
        "per_fold_test_mcc_mean": per_fold_test_mcc_mean,
        "per_fold_test_mcc_std": per_fold_test_mcc_std,
        "per_fold_threshold_mean": per_fold_threshold_mean,
        "per_fold_threshold_std": per_fold_threshold_std,
        "num_folds": int(config.split_num),
        "fold_results_path": fold_path,
    }

    # Save summary CSV (single row)
    summary_path = os.path.join(config.pred_dir, "test_results.csv")
    pd.DataFrame([result]).to_csv(summary_path, index=False)
    logger.info(f"Summary saved to: {summary_path}")

    return result


def augment_pulse_set_vsb(
    x,
    drop_rate=0.1,
    noise_std=0.01,
    scale_std=0.05,
    shuffle=False,
    min_keep=128,
):
    """
    专为 VSB 异常检测任务设计的集合级增广函数。
    x: [B, 160, 30]
    """
    B, N, D = x.shape
    device = x.device

    # 随机 drop
    keep_num = int(N * (1 - drop_rate))
    keep_num = max(keep_num, min_keep)

    mask = torch.zeros(B, N, device=device)
    for i in range(B):
        idx = torch.randperm(N)[:keep_num]
        mask[i, idx] = 1.0
    mask = mask.unsqueeze(-1)  # [B, N, 1]

    # 高斯噪声 + 缩放
    noise = torch.randn_like(x) * noise_std
    scale = 1.0 + torch.randn(B, N, 1, device=device) * scale_std
    x_aug = (x + noise) * scale

    # 应用 drop 掩码
    x_aug = x_aug * mask

    # 可选 shuffle（保证顺序无关性）
    if shuffle:
        for i in range(B):
            perm = torch.randperm(N)
            x_aug[i] = x_aug[i, perm]
            mask[i] = mask[i, perm]

    return x_aug, mask


import random


def augment_signal(
    signal,
    drop_prob=0.3,
    noise_prob=0.5,
    shuffle_prob=0.5,
    drop_fraction=0.1,
    noise_std=0.1,
    shuffle_fraction=0.3,
):
    """
    对单个信号样本应用随机增强，返回增强后的新样本。
    参数:
        signal: 张量 shape [160, 30], 单个样本的脉冲序列
        drop_prob: 应用drop token的概率
        noise_prob: 添加噪声的概率
        shuffle_prob: 扰乱顺序的概率
        drop_fraction: 丢弃脉冲的比例 (如0.1表示随机丢弃10%的脉冲)
        noise_std: 高斯噪声标准差
        shuffle_fraction: 顺序扰乱的比例 (一次打乱的脉冲片段长度占总长度的比例)
    返回:
        augmented: 张量 [160, 30], 增强后的信号
    """
    seq_len = signal.shape[0]  # 160
    augmented = signal.clone()  # 复制一份信号

    # 1. 随机丢弃部分脉冲: 将一定比例的脉冲置零
    if random.random() < drop_prob:
        drop_count = int(seq_len * drop_fraction)
        # 随机选择 drop_count 个脉冲的索引
        drop_indices = random.sample(range(seq_len), drop_count)
        augmented[drop_indices] = 0.0  # 将这些脉冲置为0

    # 2. 添加高斯噪声
    if random.random() < noise_prob:
        noise = torch.randn_like(augmented) * noise_std
        augmented = augmented + noise

    # 3. 扰乱部分脉冲顺序
    if random.random() < shuffle_prob:
        # 决定扰乱片段的长度
        seg_len = int(seq_len * shuffle_fraction)
        if seg_len > 1 and seg_len < seq_len:
            start = random.randint(0, seq_len - seg_len)
            # 提取片段并打乱顺序
            segment = augmented[start : start + seg_len].clone()
            idx = torch.randperm(seg_len)  # 生成一个随机排列
            augmented[start : start + seg_len] = segment[idx]

    return augmented, None
